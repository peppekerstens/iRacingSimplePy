from openpyxl import load_workbook
workbook = load_workbook(filename="C:\Users\peppe\OneDrive\Documenten\simracing\Trackside Endurance\PECS2\Spa24h\")
workbook.sheetnames

sheet = workbook.active
sheet




import os


#remark
currdir = os.getcwd()
workdir = "/Users/peppe/OneDrive/Documenten/simracing/Trackside Endurance/PECS2/Spa24h"
os.chdir(workdir)
obj = os.scandir()
 
# List all files and directories in the specified path
print("Files and Directories in '% s':" % workdir)
for entry in obj:
    #if entry.is_dir() or entry.is_file():
    if entry.is_file() and entry.name[-3:] == "csv":
        print(entry.name)

#get a file list



import csv
with open('eventresult_50906798.csv', newline='') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=',', quotechar='"')
    for row in spamreader:
        print(', '.join(row))